from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import time

class Attendance():
    def __init__(self,att,subject,sub_topic,core,hour):
        self.att = att
        self.subject = subject
        self.sub_topic = sub_topic
        self.core = core
        self.hour = hour

class mark_attendance(Attendance):
    def __init__(self,att,subject,sub_topic,core,hour):
        super().__init__(att,subject,sub_topic,core,hour)
        workbook = load_workbook("Attendance Tracking.xlsx")
        sheet = workbook["Sheet1"]

        green_fill = PatternFill(start_color="FF4CAF50", end_color="FF4CAF50", fill_type="solid")
        orange_fill = PatternFill(start_color="FFEF6C00", end_color="FFEF6C00", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")





        date = str(datetime.now()).split()

        
        present_date = date[0] #present date
        col_num = 4


        for row in range(6,sheet.max_row+1):

            cell_raw = sheet.cell(row=row,column=col_num).value
            if cell_raw is not None:

                cell_value = str(cell_raw).split()
                #stoped to present date
                
                if cell_value[0] == present_date:
                    now = datetime.now()
                    current_time = now.strftime("%H:%M") # current time on Based 24 Hour
                    col_num+=1

                    for r in range(row,sheet.max_row+1):
                        timee = str(sheet.cell(row=r,column = col_num).value)

                        if timee is not None:
                            time_str = str(timee).strip()

                            if time_str == "Total:":
                                break
                            else:
                                t = time_str.split('-')
                                start = t[0].strip()
                                ending = t[1].strip()

                                fmt = "%H:%M"

                                current = datetime.strptime(current_time, fmt).time()
                                start_time = datetime.strptime(start, fmt).time()
                                end_time = datetime.strptime(ending, fmt).time()
                                print(current)
                    
                                if start_time <= current <= end_time:
                                    #add value on the base of timeline


                                    col_num+=1
                                    


                                    if self.att == "P":
                                        #insert & update

                                        sheet.cell(row=r,column=col_num).value = self.att
                                        sheet.cell(row=r,column=col_num).fill = green_fill
                                        col_num+=1

                                        sheet.cell(row=r,column=col_num).value = self.subject
                                        sheet.cell(row=r,column=col_num).fill = white_fill
                                        col_num+=1

                                        sheet.cell(row=r,column=col_num).value = self.sub_topic
                                        sheet.cell(row=r,column=col_num).fill = white_fill
                                        col_num+=1

                                        sheet.cell(row=r,column=col_num).value = self.core
                                        sheet.cell(row=r,column=col_num).fill = white_fill

                                        col_num+=1
                                        sheet.cell(row=r,column=col_num).value = self.hour

                                        print("")
                                    
                                    else:
                                        if self.att == "A":
                                                
                                            #modify
                                            sheet.cell(row=r,column=col_num).value = self.att
                                            sheet.cell(row=r,column=col_num).fill = orange_fill

                                            col_num+=1
                                            sheet.cell(row=r,column=col_num).value = self.subject
                                            sheet.cell(row=r,column=col_num).fill = orange_fill
                                            col_num+=1
                                            sheet.cell(row=r,column=col_num).value = self.sub_topic
                                            sheet.cell(row=r,column=col_num).fill = orange_fill
                                            col_num+=1
                                            sheet.cell(row=r,column=col_num).value = self.core
                                            sheet.cell(row=r,column=col_num).fill = orange_fill
                                            
                                            col_num+=1
                                            sheet.cell(row=r,column=col_num).value = self.hour
                                            print("You Absent Today")

        
                                    
                                    
                                    


                                    print("\n.",end="")
                                    time.sleep(1)
                                    print(".",end="")
                                    time.sleep(0.6)
                                    print(".Attendance Marked>>>")
                                    break
            
            
            
        workbook.save("Attendance Tracking.xlsx")
                                    


print("Attendance Update==>\n")                                
while True:                                
    print("Enter Present For (P) / Abent For (A)")
    att = input(">>").strip().upper()
    if att not in ["P","A"]:
        print("You Entered Invalid Details")
        continue
    else:
        break
    
print()
if att == "P":
    print("Enter Subject Name")
    sub_name = input(">>").title()
    print()
    print("Enter Topic Name")
    topic = input(">>").title()
    print()
    print("Enter Sub-Topic Name")
    sub_topic = input(">>").title()
    print(f"\nTell me Exact Time that spend on {sub_name} Subject?")
    hours = input(">>")
    hours = datetime.strptime(hours, "%H:%M").time()
    

else:
    #by default for Full Row Absent 
    sub_name = topic = sub_topic = "Absent"
    hours = "0:00"
    



d = mark_attendance(att,sub_name,topic,sub_topic,hours)
